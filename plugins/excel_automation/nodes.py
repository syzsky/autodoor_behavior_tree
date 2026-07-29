# plugins/excel_automation/nodes.py
"""Excel 自动化节点"""
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from bt_core.nodes import ActionNode
from bt_core.status import NodeStatus


class ExcelWriteNode(ActionNode):
    """Excel 写入节点 — 将黑板中的二维数据写入 Excel 文件"""

    NODE_TYPE = "ExcelWriteNode"

    def _execute_action(self, context):
        if not OPENPYXL_AVAILABLE:
            return NodeStatus.FAILURE

        file_path = self.config.get("file_path", "")
        sheet_name = self.config.get("sheet_name", "Sheet1")
        data_key = self.config.get("data_key", "table_data")
        start_cell = self.config.get("start_cell", "A1")

        if not file_path:
            return NodeStatus.FAILURE

        data = context.blackboard.get(data_key)
        if not data or not isinstance(data, (list, tuple)):
            return NodeStatus.FAILURE

        try:
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name

            # 解析起始单元格
            from openpyxl.utils import range_boundaries
            col, row, _, _ = range_boundaries(start_cell + ":" + start_cell)

            for r_idx, row_data in enumerate(data):
                for c_idx, value in enumerate(row_data):
                    ws.cell(row=row + r_idx, column=col + c_idx, value=value)

            wb.save(file_path)
            return NodeStatus.SUCCESS
        except Exception:
            return NodeStatus.FAILURE


class ExcelReadNode(ActionNode):
    """Excel 读取节点 — 读取 Excel 范围到黑板"""

    NODE_TYPE = "ExcelReadNode"

    def _execute_action(self, context):
        if not OPENPYXL_AVAILABLE:
            return NodeStatus.FAILURE

        file_path = self.config.get("file_path", "")
        sheet_name = self.config.get("sheet_name", "")
        cell_range = self.config.get("cell_range", "A1:Z100")
        target_key = self.config.get("target_key", "excel_data")

        if not file_path or not os.path.exists(file_path):
            return NodeStatus.FAILURE

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 读取范围
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)

            data = []
            for r in range(min_row, max_row + 1):
                row = []
                for c in range(min_col, max_col + 1):
                    row.append(ws.cell(row=r, column=c).value)
                data.append(row)

            context.blackboard.set(target_key, data)
            return NodeStatus.SUCCESS
        except Exception:
            return NodeStatus.FAILURE


class ExcelFormatNode(ActionNode):
    """Excel 格式化节点 — 应用单元格格式"""

    NODE_TYPE = "ExcelFormatNode"

    def _execute_action(self, context):
        if not OPENPYXL_AVAILABLE:
            return NodeStatus.FAILURE

        file_path = self.config.get("file_path", "")
        sheet_name = self.config.get("sheet_name", "")
        cell_range = self.config.get("cell_range", "A1:A1")
        bold = self.config.get("bold", False)
        bg_color = self.config.get("bg_color", "")

        if not file_path or not os.path.exists(file_path):
            return NodeStatus.FAILURE

        try:
            wb = openpyxl.load_workbook(file_path)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            font = Font(bold=bold)
            fill = PatternFill(start_color=bg_color, end_color=bg_color,
                               fill_type="solid") if bg_color else None

            for row in ws[cell_range]:
                for cell in row:
                    cell.font = font
                    if fill:
                        cell.fill = fill

            wb.save(file_path)
            return NodeStatus.SUCCESS
        except Exception:
            return NodeStatus.FAILURE
